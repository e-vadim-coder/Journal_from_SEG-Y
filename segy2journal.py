#!/usr/bin/env python3
"""
SEG-Y to Journal (XLSX)
- Сканирует каталог на SEG-Y файлы
- Если в имени есть дата (XXXXX_YYYY_MM_DD_NNN) — сортирует по дате и номеру
- Извлекает дату/время начала и конца записи, координаты начала и конца
- Сохраняет журнал в XLSX
"""

import os
import re
import sys
import struct
from datetime import datetime, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# ─── SEG-Y reader ───────────────────────────────────────────────────────────

SEGY_BIN_HDR_SIZE = 400       # binary file header
SEGY_TRACE_HDR_SIZE = 240     # trace header
SEGY_TEXT_HDR_SIZE = 3200     # textual file header


def read_segy_text_header(f):
    """Read 3200-byte EBCDIC/ASCII text header."""
    raw = f.read(SEGY_TEXT_HDR_SIZE)
    if len(raw) < SEGY_TEXT_HDR_SIZE:
        return ""
    # Try ASCII first, then EBCDIC
    try:
        return raw.decode("ascii", errors="replace")
    except Exception:
        try:
            return raw.decode("cp500", errors="replace")
        except Exception:
            return raw.decode("latin-1", errors="replace")


def read_segy_bin_header(f, endian=">"):
    """Read binary file header (400 bytes), return dict of key fields."""
    raw = f.read(SEGY_BIN_HDR_SIZE)
    if len(raw) < SEGY_BIN_HDR_SIZE:
        return {}
    fmt = endian
    h = {}
    # Job identification
    h["job_id"] = struct.unpack(f"{fmt}i", raw[0:4])[0]
    h["line_num"] = struct.unpack(f"{fmt}i", raw[4:8])[0]
    h["reel_num"] = struct.unpack(f"{fmt}i", raw[8:12])[0]
    # Number of traces per record
    h["ntr_per_record"] = struct.unpack(f"{fmt}h", raw[12:14])[0]
    # Number of auxiliary traces
    h["aux_traces"] = struct.unpack(f"{fmt}h", raw[14:16])[0]
    # Sample interval (microseconds)
    h["sample_int"] = struct.unpack(f"{fmt}h", raw[16:18])[0]
    # Sample interval original
    h["sample_int_orig"] = struct.unpack(f"{fmt}h", raw[18:20])[0]
    # Number of samples per trace
    h["ns_per_trace"] = struct.unpack(f"{fmt}h", raw[20:22])[0]
    # Number of samples per trace original
    h["ns_per_trace_orig"] = struct.unpack(f"{fmt}h", raw[22:24])[0]
    # Data sample format code
    h["format_code"] = struct.unpack(f"{fmt}h", raw[24:26])[0]
    # Measurement system
    h["measure_sys"] = struct.unpack(f"{fmt}h", raw[54:56])[0]
    # Rev 1: bytes 115-116 sample interval
    h["sample_int_rev1"] = struct.unpack(f"{fmt}h", raw[115:117])[0]
    # Rev 1: bytes 117-118 sample interval original
    h["sample_int_orig_rev1"] = struct.unpack(f"{fmt}h", raw[117:119])[0]
    # Recording date/time in binary header (bytes 321-324)
    h["rec_year"] = struct.unpack(f"{fmt}h", raw[321:323])[0]
    h["rec_day"] = struct.unpack(f"{fmt}h", raw[323:325])[0]
    return h


def parse_trace_header(raw, endian=">"):
    """Parse 240-byte trace header, return dict."""
    if len(raw) < SEGY_TRACE_HDR_SIZE:
        return None
    fmt = endian
    th = {}
    th["seq_line"] = struct.unpack(f"{fmt}i", raw[0:4])[0]
    th["seq_file"] = struct.unpack(f"{fmt}i", raw[4:8])[0]
    th["field_rec"] = struct.unpack(f"{fmt}i", raw[8:12])[0]
    th["trace_in_rec"] = struct.unpack(f"{fmt}i", raw[12:16])[0]
    th["src_point"] = struct.unpack(f"{fmt}i", raw[16:20])[0]
    th["cdp_ens"] = struct.unpack(f"{fmt}i", raw[20:24])[0]
    th["cdp_trace"] = struct.unpack(f"{fmt}i", raw[24:28])[0]
    th["trace_id"] = struct.unpack(f"{fmt}h", raw[28:30])[0]
    # Coordinate scalar (bytes 69-70 1-based → 0-based 68:70)
    th["coord_scalar"] = struct.unpack(f"{fmt}h", raw[68:70])[0]
    # Source coordinates (bytes 73-76, 77-80 1-based → 0-based 72:76, 76:80)
    th["src_x"] = struct.unpack(f"{fmt}i", raw[72:76])[0]
    th["src_y"] = struct.unpack(f"{fmt}i", raw[76:80])[0]
    # Group coordinates (bytes 81-88) — не используем в журнале
    # CDP coordinates (bytes 181-184, 185-188 1-based → 0-based 180:184, 184:188)
    th["cdp_x"] = struct.unpack(f"{fmt}i", raw[180:184])[0]
    th["cdp_y"] = struct.unpack(f"{fmt}i", raw[184:188])[0]

    # Apply scalar:
    #   scalar > 0  → multiply by scalar
    #   scalar < 0  → divide by |scalar|
    #   scalar == 0 → no scaling
    s = th["coord_scalar"]
    for key in ("src_x", "src_y", "cdp_x", "cdp_y"):
        val = th[key]
        if s != 0 and val != 0:
            if s > 0:
                val = val * s
            else:
                val = val / (-s)
        # Source coords are in seconds → convert to degrees
        if key.startswith("src_"):
            val = val / 3600.0
        th[key] = val

    # Date/time (bytes 157-166 in 1-based SEG-Y → 0-based indices 156-166)
    th["yr"] = struct.unpack(f"{fmt}h", raw[156:158])[0]
    th["day_of_year"] = struct.unpack(f"{fmt}h", raw[158:160])[0]
    th["hour"] = struct.unpack(f"{fmt}h", raw[160:162])[0]
    th["minute"] = struct.unpack(f"{fmt}h", raw[162:164])[0]
    th["second"] = struct.unpack(f"{fmt}h", raw[164:166])[0]
    th["time_base"] = struct.unpack(f"{fmt}h", raw[207:209])[0]

    # Trace start time (delay recording time, bytes 109-112)
    th["delay_ms"] = struct.unpack(f"{fmt}h", raw[109:111])[0]

    return th


def day_of_year_to_date(year, doy):
    """Convert day-of-year to datetime.date."""
    try:
        return datetime(year, 1, 1) + timedelta(days=int(doy) - 1)
    except Exception:
        return None


def trace_time_to_dt(th):
    """Convert trace header date/time fields to datetime.
    Returns (dt, raw_yr, raw_doy) — dt=None if conversion fails."""
    yr = th.get("yr", 0)
    doy = th.get("day_of_year", 0)
    hr = th.get("hour", 0)
    mn = th.get("minute", 0)
    sc = th.get("second", 0)
    if yr == 0:
        return None, yr, doy
    d = day_of_year_to_date(yr, doy)
    if d is None:
        return None, yr, doy
    try:
        return d.replace(hour=int(hr), minute=int(mn), second=int(sc)), yr, doy
    except Exception:
        return d, yr, doy


def apply_coord_scalar(val, scalar):
    if scalar is None or scalar == 0:
        return val
    if scalar > 0:
        return val / scalar
    return -val * scalar


def probe_trace_endian(fpath, bin_endian=">"):
    """Probe endianness of trace headers separately from binary header.
    The binary header of SEG-Y is always big-endian, but trace headers
    can be little-endian in some implementations.
    Uses time fields (year, day, hour, minute, second) as strongest signal.
    """
    with open(fpath, "rb") as f:
        f.read(SEGY_TEXT_HDR_SIZE + SEGY_BIN_HDR_SIZE)
        th = f.read(SEGY_TRACE_HDR_SIZE)
        if len(th) < SEGY_TRACE_HDR_SIZE:
            return bin_endian

        def score(end):
            s = 0
            # Year (0-based 156-158) — strongest signal
            yr = struct.unpack(f"{end}h", th[156:158])[0]
            if 1900 <= yr <= 2100:
                s += 30
            # Day of year (158-160)
            doy = struct.unpack(f"{end}h", th[158:160])[0]
            if 1 <= doy <= 366:
                s += 10
            # Hour (160-162)
            hr = struct.unpack(f"{end}h", th[160:162])[0]
            if 0 <= hr <= 23:
                s += 5
            # Minute (162-164)
            mn = struct.unpack(f"{end}h", th[162:164])[0]
            if 0 <= mn <= 59:
                s += 5
            # Second (164-166)
            sc = struct.unpack(f"{end}h", th[164:166])[0]
            if 0 <= sc <= 59:
                s += 5
            # seq_file (bytes 5-8 1-based → 0-based 4:8)
            sf = struct.unpack(f"{end}i", th[4:8])[0]
            if sf == 1:
                s += 15
            # Scalar (bytes 69-70 1-based → 0-based 68:70)
            sc_val = struct.unpack(f"{end}h", th[68:70])[0]
            if sc_val != 0 and abs(sc_val) in (1, 10, 100, 1000, 10000, 100000, 1000000):
                s += 15
            # Coordinate plausibility (bytes 73-76, 77-80 1-based → 0-based 72:76, 76:80)
            src_x = struct.unpack(f"{end}i", th[72:76])[0]
            src_y = struct.unpack(f"{end}i", th[76:80])[0]
            if 0 < src_x < 100000000 and 0 < src_y < 100000000:
                s += 5
            return s

        s_be = score(">")
        s_le = score("<")
        return "<" if s_le > s_be else ">"


def process_segy(fpath):
    """
    Open SEG-Y file, extract metadata.
    Returns dict with keys: filename, file_date, file_num,
    start_dt, end_dt, src_lon_start, src_lat_start, src_lon_end, src_lat_end,
    cdp_x_start, cdp_y_start, cdp_x_end, cdp_y_end,
    ns_per_trace, sample_int_us, n_traces, format_code, notes, anomalies
    """
    result = {
        "filename": os.path.basename(fpath),
        "filepath": fpath,
        "start_dt": None,
        "end_dt": None,
        "_start_raw": None, "_end_raw": None,
        "src_lon_start": None, "src_lat_start": None,
        "src_lon_end": None, "src_lat_end": None,
        "cdp_x_start": None, "cdp_y_start": None,
        "cdp_x_end": None, "cdp_y_end": None,
        "ns_per_trace": None, "sample_int_us": None,
        "n_traces": 0, "format_code": None,
        "file_date": None, "file_num": None,
        "notes": "",
        "anomalies": [],
    }

    # Binary header is always big-endian per SEG-Y standard
    with open(fpath, "rb") as f:
        text_hdr = read_segy_text_header(f)
        bin_hdr = read_segy_bin_header(f, ">")
        if not bin_hdr:
            result["notes"] = "Cannot read binary header"
            return result

    # Trace headers may have different endianness — probe from first trace
    trace_endian = probe_trace_endian(fpath, ">")

    with open(fpath, "rb") as f:
        f.read(SEGY_TEXT_HDR_SIZE + SEGY_BIN_HDR_SIZE)

        result["ns_per_trace"] = bin_hdr.get("ns_per_trace", 0)
        result["sample_int_us"] = bin_hdr.get("sample_int", 0)
        result["format_code"] = bin_hdr.get("format_code", 0)

        if result["ns_per_trace"] == 0:
            result["ns_per_trace"] = bin_hdr.get("ns_per_trace_orig", 0)
        if result["sample_int_us"] == 0:
            result["sample_int_us"] = bin_hdr.get("sample_int_rev1", 0)
            if result["sample_int_us"] == 0:
                result["sample_int_us"] = bin_hdr.get("sample_int_orig_rev1", 0)

        # Compute trace data size
        ns = result["ns_per_trace"] or 0
        fmt_code = result["format_code"] or 1
        sample_bytes = {1: 4, 2: 4, 3: 2, 4: 4, 5: 4, 6: 4, 7: 3, 8: 1}
        bps = sample_bytes.get(fmt_code, 4)
        trace_data_size = ns * bps

        # Count traces from file size
        f.seek(0, os.SEEK_END)
        file_size = f.tell()
        trace_total_size = SEGY_TRACE_HDR_SIZE + trace_data_size
        data_start = SEGY_TEXT_HDR_SIZE + SEGY_BIN_HDR_SIZE
        n_traces = (file_size - data_start) // trace_total_size
        result["n_traces"] = n_traces

        # Read first trace header
        f.seek(data_start)
        raw = f.read(SEGY_TRACE_HDR_SIZE)
        first_th = parse_trace_header(raw, trace_endian) if len(raw) >= SEGY_TRACE_HDR_SIZE else None

        # Read last trace header
        if n_traces > 1:
            last_offset = data_start + (n_traces - 1) * trace_total_size
            f.seek(last_offset)
            raw = f.read(SEGY_TRACE_HDR_SIZE)
            last_th = parse_trace_header(raw, trace_endian) if len(raw) >= SEGY_TRACE_HDR_SIZE else None
        else:
            last_th = first_th

        result["n_traces"] = n_traces

        # Start/end datetimes from first/last trace headers
        if first_th:
            start_dt, start_yr, start_doy = trace_time_to_dt(first_th)
            result["start_dt"] = start_dt
            result["_start_raw"] = (start_yr, start_doy)
            result["src_lon_start"] = first_th.get("src_x")
            result["src_lat_start"] = first_th.get("src_y")
            result["cdp_x_start"] = first_th.get("cdp_x")
            result["cdp_y_start"] = first_th.get("cdp_y")

        if last_th:
            end_dt, end_yr, end_doy = trace_time_to_dt(last_th)
            result["end_dt"] = end_dt
            result["_end_raw"] = (end_yr, end_doy)
            result["src_lon_end"] = last_th.get("src_x")
            result["src_lat_end"] = last_th.get("src_y")
            result["cdp_x_end"] = last_th.get("cdp_x")
            result["cdp_y_end"] = last_th.get("cdp_y")

    # Fallback: use binary header date if no trace time found
    if result["start_dt"] is None:
        yr = bin_hdr.get("rec_year", 0)
        dy = bin_hdr.get("rec_day", 0)
        if yr > 0:
            result["start_dt"] = day_of_year_to_date(yr, dy)
            result["notes"] = "Date from binary header (no trace time)"

    # Detect anomalies
    detect_anomalies(result)

    return result


# ─── Anomaly detection ──────────────────────────────────────────────────────

def detect_anomalies(result):
    """Check extracted data for anomalies, populate result['anomalies'] list."""
    anomalies = []
    notes = result.get("notes", "") or ""

    # Time anomalies: distinguish "zero in file" from "bad read/conversion"
    for field, label, raw_key in [
        ("start_dt", "Время начала записи", "_start_raw"),
        ("end_dt", "Время конца записи", "_end_raw"),
    ]:
        dt = result.get(field)
        raw = result.get(raw_key)
        if dt is None:
            if raw and raw[0] == 0:
                anomalies.append(field)
                notes += f"{label} = 0 в заголовках трасс; "
            elif raw and raw[0] != 0:
                anomalies.append(field)
                notes += f"{label}: прочитаны некорректные значения (год={raw[0]}, день={raw[1]}); "
            else:
                anomalies.append(field)
                notes += f"{label} не удалось прочитать; "

    # Coordinate bounds: lon [-180, 180], lat [-90, 90]
    for field, label in [
        ("src_lon_start", "долгота нач."),
        ("src_lat_start", "широта нач."),
        ("src_lon_end", "долгота кон."),
        ("src_lat_end", "широта кон."),
    ]:
        val = result.get(field)
        if val is not None:
            v = float(val)
            if field.startswith("src_lon") and not (-180 <= v <= 180):
                anomalies.append(field)
                notes += f"Координата {label} ({v:.4f}°) вне допустимого диапазона; "
            elif field.startswith("src_lat") and not (-90 <= v <= 90):
                anomalies.append(field)
                notes += f"Координата {label} ({v:.4f}°) вне допустимого диапазона; "

    # CDP coordinate zeros
    for field, label in [
        ("cdp_x_start", "CDP X нач."),
        ("cdp_y_start", "CDP Y нач."),
        ("cdp_x_end", "CDP X кон."),
        ("cdp_y_end", "CDP Y кон."),
    ]:
        val = result.get(field)
        if val is not None and float(val) == 0:
            anomalies.append(field)
            notes += f"Координата {label} = 0; "

    # CDP coordinate suspicious range (if absolute value > 10^7 — likely bad)
    for field, label in [
        ("cdp_x_start", "CDP X нач."),
        ("cdp_y_start", "CDP Y нач."),
        ("cdp_x_end", "CDP X кон."),
        ("cdp_y_end", "CDP Y кон."),
    ]:
        val = result.get(field)
        if val is not None:
            v = float(val)
            if abs(v) > 10_000_000:
                anomalies.append(field)
                notes += f"Координата {label} ({v:.0f} м) аномально велика; "

    result["anomalies"] = anomalies
    result["notes"] = notes.strip().rstrip(";")


# ─── Filename date parser ───────────────────────────────────────────────────

def parse_filename_date(fname):
    """
    Try to extract date and number from filename.
    Supported patterns:
      XXXXX_YYYY_MM_DD_NNN  (or with - or . separators)
      XXXXX_YYMMDD_NN       (2-digit year: 20YY)
      YYYYMMDD or YYMMDD at start or in the name
    Returns (dt, num) or (None, None)
    """
    stem = os.path.splitext(fname)[0]

    # Pattern 1: XXXXX_YYYY_MM_DD_NNN
    m = re.search(r'_(\d{4})[_\-\.](\d{2})[_\-\.](\d{2})[_\-\.](\d+)', stem)
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            num = int(m.group(4))
            return dt, num
        except ValueError:
            pass

    # Pattern 1b: XXXXX_YYMMDD_NN  (2-digit year)
    m = re.search(r'_(\d{2})(\d{2})(\d{2})[_\-\.](\d+)', stem)
    if m:
        try:
            yy = int(m.group(1))
            year = 2000 + yy if yy < 50 else 1900 + yy
            dt = datetime(year, int(m.group(2)), int(m.group(3)))
            num = int(m.group(4))
            return dt, num
        except ValueError:
            pass

    # Pattern 2: YYYY_MM_DD anywhere
    m = re.search(r'(\d{4})[_\-\.](\d{2})[_\-\.](\d{2})', stem)
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            # Look for a trailing number
            mm = re.search(r'_(\d+)$', stem.replace("-", "_").replace(".", "_"))
            num = int(mm.group(1)) if mm else 0
            return dt, num
        except ValueError:
            pass

    # Pattern 3: YYYYMMDD at start
    m = re.match(r'(\d{4})(\d{2})(\d{2})', stem)
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            mm = re.search(r'_(\d+)$', stem.replace("-", "_").replace(".", "_"))
            num = int(mm.group(1)) if mm else 0
            return dt, num
        except ValueError:
            pass

    # Pattern 4: DD_MM_YYYY or MM_DD_YYYY (european-style)
    # YYYY is last 4-digit group
    m = re.search(r'(\d{2})[_\-\.](\d{2})[_\-\.](\d{4})', stem)
    if m:
        try:
            # Try as DD_MM_YYYY
            dt = datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            mm = re.search(r'_(\d+)$', stem.replace("-", "_").replace(".", "_"))
            num = int(mm.group(1)) if mm else 0
            return dt, num
        except ValueError:
            pass

    return None, None


# ─── Main processing ────────────────────────────────────────────────────────

def scan_directory(root_dir, recursive=True):
    """Find all SEG-Y files in directory."""
    segy_exts = {".sgy", ".segy", ".SEGY", ".SGY"}
    files = []
    for dirpath, dirnames, filenames in os.walk(root_dir):
        if not recursive:
            dirnames.clear()
        for fname in filenames:
            ext = os.path.splitext(fname)[1]
            if ext.lower() in (".sgy", ".segy"):
                files.append(os.path.join(dirpath, fname))
    return files


def process_directory(files, root_dir, recursive=True):
    """Process all SEG-Y files in directory, return sorted results."""
    if not files:
        print(f"No SEG-Y files found in {root_dir}")
        return []

    total = len(files)
    print(f"Found {total} SEG-Y file(s), processing...\n")

    results = []
    no_date = []

    for i, fpath in enumerate(files, 1):
        fname = os.path.basename(fpath)
        dt, num = parse_filename_date(fname)

        result = process_segy(fpath)
        result["file_date"] = dt
        result["file_num"] = num

        if dt:
            results.append(result)
        else:
            no_date.append(result)

        status = f"[{i}/{total}]  {fname}"
        if dt:
            status += f"  [{dt.date()} #{num}]"
        if result["n_traces"]:
            status += f"  {result['n_traces']} traces"
        if result["start_dt"]:
            status += f"  start={result['start_dt']}"
        print(status)

    # Sort: date ascending, then number ascending
    results.sort(key=lambda r: (r["file_date"] or datetime.min, r.get("file_num", 0) or 0))

    # Append files without date at the end (sorted by filename)
    no_date.sort(key=lambda r: r["filename"])
    results.extend(no_date)

    return results


# ─── XLSX writer ────────────────────────────────────────────────────────────

def write_journal_xlsx(results, output_path):
    """Write the journal table to XLSX."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал SEG-Y"

    # Styles
    hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="center", wrap_text=False)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    anomaly_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
    anomaly_font = Font(name="Calibri", size=10, color="CC0000")

    # Headers (RU)
    headers = [
        "№ п/п",
        "Файл",
        "Дата (из имени)",
        "Номер (из имени)",
        "Начало записи",
        "Конец записи",
        "Кол-во трасс",
        "Длина трассы (отсчётов)",
        "Интервал дискр. (мкс)",
        "Долгота WGS84 (°) — нач.",
        "Широта WGS84 (°) — нач.",
        "Долгота WGS84 (°) — кон.",
        "Широта WGS84 (°) — кон.",
        "X (CDP) — нач., м",
        "Y (CDP) — нач., м",
        "X (CDP) — кон., м",
        "Y (CDP) — кон., м",
        "Примечания",
    ]

    # Write headers
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Column widths
    col_widths = {
        1: 7,    # №
        2: 45,   # Файл
        3: 14,   # Дата
        4: 10,   # Номер
        5: 18,   # Начало
        6: 18,   # Конец
        7: 10,   # Трассы
        8: 12,   # Длина
        9: 12,   # Интервал
        10: 16,  # Долгота нач.
        11: 16,  # Широта нач.
        12: 16,  # Долгота кон.
        13: 16,  # Широта кон.
        14: 13,  # X CDP нач.
        15: 13,  # Y CDP нач.
        16: 13,  # X CDP кон.
        17: 13,  # Y CDP кон.
        18: 55,  # Notes
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Mapping from anomaly field name → column number
    ANOMALY_COLS = {
        "start_dt": 5, "end_dt": 6,
        "src_lon_start": 10, "src_lat_start": 11,
        "src_lon_end": 12, "src_lat_end": 13,
        "cdp_x_start": 14, "cdp_y_start": 15,
        "cdp_x_end": 16, "cdp_y_end": 17,
    }

    # Write data
    for i, r in enumerate(results):
        row = i + 2
        anomalies = r.get("anomalies", [])
        has_anomalies = len(anomalies) > 0

        ws.cell(row=row, column=1, value=i + 1).font = cell_font
        ws.cell(row=row, column=1).alignment = center_align

        ws.cell(row=row, column=2, value=r["filename"]).font = cell_font

        # Date from filename
        if r["file_date"]:
            ws.cell(row=row, column=3, value=r["file_date"]).font = cell_font
            ws.cell(row=row, column=3).number_format = "DD.MM.YYYY"
        else:
            ws.cell(row=row, column=3, value="—").font = cell_font

        # Number
        ws.cell(row=row, column=4, value=r.get("file_num", "—")).font = cell_font
        ws.cell(row=row, column=4).alignment = center_align

        # Start time
        if r["start_dt"]:
            ws.cell(row=row, column=5, value=r["start_dt"]).font = cell_font
            ws.cell(row=row, column=5).number_format = "DD.MM.YYYY HH:MM:SS"
        else:
            ws.cell(row=row, column=5, value="—").font = cell_font

        # End time
        if r["end_dt"]:
            ws.cell(row=row, column=6, value=r["end_dt"]).font = cell_font
            ws.cell(row=row, column=6).number_format = "DD.MM.YYYY HH:MM:SS"
        else:
            ws.cell(row=row, column=6, value="—").font = cell_font

        # Numeric fields
        ws.cell(row=row, column=7, value=r.get("n_traces", 0)).font = cell_font
        ws.cell(row=row, column=7).alignment = center_align

        ws.cell(row=row, column=8, value=r.get("ns_per_trace", "—")).font = cell_font
        ws.cell(row=row, column=8).alignment = center_align

        ws.cell(row=row, column=9, value=r.get("sample_int_us", "—")).font = cell_font
        ws.cell(row=row, column=9).alignment = center_align

        # Coordinates:
        #   cols 10-13: source lon/lat in WGS84 degrees (start, end)
        #   cols 14-17: CDP X/Y in metres (start, end)
        def write_deg(ws, row, col, val):
            if val is not None:
                cell = ws.cell(row=row, column=col, value=float(val))
                cell.font = cell_font
                cell.number_format = "0.000000"
            else:
                ws.cell(row=row, column=col, value="—").font = cell_font

        def write_metric(ws, row, col, val):
            if val is not None:
                cell = ws.cell(row=row, column=col, value=float(val))
                cell.font = cell_font
                cell.number_format = "0.0########"
            else:
                ws.cell(row=row, column=col, value="—").font = cell_font

        write_deg(ws, row, 10, r.get("src_lon_start"))
        write_deg(ws, row, 11, r.get("src_lat_start"))
        write_deg(ws, row, 12, r.get("src_lon_end"))
        write_deg(ws, row, 13, r.get("src_lat_end"))
        write_metric(ws, row, 14, r.get("cdp_x_start"))
        write_metric(ws, row, 15, r.get("cdp_y_start"))
        write_metric(ws, row, 16, r.get("cdp_x_end"))
        write_metric(ws, row, 17, r.get("cdp_y_end"))

        # Notes — combine with anomaly info
        notes = r.get("notes", "")
        if not notes and r["file_date"] is None and r["start_dt"] is None:
            notes = "Дата не определена"
        ws.cell(row=row, column=18, value=notes).font = cell_font

        # Red highlight for anomalous cells
        if has_anomalies:
            for field in anomalies:
                col = ANOMALY_COLS.get(field)
                if col:
                    ws.cell(row=row, column=col).fill = anomaly_fill
                    ws.cell(row=row, column=col).font = anomaly_font
            # Also highlight the notes column
            ws.cell(row=row, column=18).fill = anomaly_fill
            ws.cell(row=row, column=18).font = anomaly_font

        # Alternate row coloring (respect existing fills)
        if not has_anomalies and i % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = alt_fill

        # Borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"

    # Try saving with retry on PermissionError
    final_path = output_path
    saved = False
    for attempt in range(100):
        try:
            wb.save(final_path)
            saved = True
            break
        except PermissionError:
            if attempt == 0:
                base, ext = os.path.splitext(final_path)
                final_path = f"{base}_1{ext}"
            else:
                base = re.sub(r'_\d+$', '', os.path.splitext(final_path)[0])
                ext = os.path.splitext(final_path)[1]
                n = int(os.path.splitext(final_path)[0].split('_')[-1]) + 1
                final_path = f"{base}_{n}{ext}"

    if saved:
        print(f"\nJournal saved: {final_path}")
    else:
        print(f"\nError: cannot write file — permission denied after 100 attempts")
    print(f"Total entries: {len(results)}")

    # Write protocol file
    write_protocol_txt(results, final_path)


# ─── Protocol file ───────────────────────────────────────────────────────────

def write_protocol_txt(results, journal_path):
    """Write a text protocol file listing all anomalies."""
    anomalous = [r for r in results if r.get("anomalies")]
    if not anomalous:
        return

    proto_path = os.path.splitext(journal_path)[0] + "_protocol.txt"
    with open(proto_path, "w", encoding="utf-8-sig") as f:
        f.write("ПРОТОКОЛ АНОМАЛИЙ SEG-Y\n")
        f.write("=" * 70 + "\n")
        f.write(f"Всего файлов: {len(results)}\n")
        f.write(f"Файлов с аномалиями: {len(anomalous)}\n")
        f.write(f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n")
        f.write("=" * 70 + "\n\n")

        for r in anomalous:
            f.write(f"Файл: {r['filename']}\n")
            f.write(f"  Аномальные поля: {', '.join(r['anomalies'])}\n")
            f.write(f"  Примечание: {r.get('notes', '')}\n")

            # Detail each anomaly
            for field in r["anomalies"]:
                labels = {
                    "start_dt": "Время начала записи",
                    "end_dt": "Время конца записи",
                    "src_lon_start": "Долгота WGS84 (°) — нач.",
                    "src_lat_start": "Широта WGS84 (°) — нач.",
                    "src_lon_end": "Долгота WGS84 (°) — кон.",
                    "src_lat_end": "Широта WGS84 (°) — кон.",
                    "cdp_x_start": "X (CDP) — нач., м",
                    "cdp_y_start": "Y (CDP) — нач., м",
                    "cdp_x_end": "X (CDP) — кон., м",
                    "cdp_y_end": "Y (CDP) — кон., м",
                }
                label = labels.get(field, field)
                value = r.get(field)
                val_str = f"{value}" if value is not None else "—"
                f.write(f"    {label} = {val_str}\n")
            f.write("\n")

    print(f"Protocol saved: {proto_path}")


# ─── CLI entry point ────────────────────────────────────────────────────────

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="SEG-Y → Journal XLSX. Извлекает даты, время и координаты из SEG-Y файлов."
    )
    parser.add_argument(
        "directory",
        nargs="?",
        default=None,
        help="Directory with SEG-Y files (if omitted — interactive dialog)",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="Output XLSX file (default: segy_journal.xlsx in the selected dir)",
    )
    parser.add_argument(
        "--no-recursive",
        action="store_false",
        dest="recursive",
        default=True,
        help="Do not scan subdirectories",
    )
    parser.add_argument(
        "--list-formats",
        action="store_true",
        help="List SEG-Y data format codes and exit",
    )

    args = parser.parse_args()

    if args.list_formats:
        print("SEG-Y Data Sample Format Codes:")
        for k, v in {1: "IBM Float (4 байта)", 2: "32-bit Integer", 3: "16-bit Integer",
                      4: "IEEE Float (4 байта)", 5: "IEEE Float (Rev 1)",
                      6: "64-bit Integer", 7: "24-bit Integer", 8: "8-bit Integer"}.items():
            print(f"  {k}: {v}")
        return

    # Choose directory (file dialog so user can see files)
    root = args.directory
    if root is None:
        try:
            import tkinter as tk
            from tkinter import filedialog
            tk_root = tk.Tk()
            tk_root.withdraw()
            file_selection = filedialog.askopenfilenames(
                title="Выберите файл SEG-Y из нужного каталога (будет обработан весь каталог)",
                filetypes=[("SEG-Y files", "*.sgy *.segy *.SGY *.SEGY"), ("All files", "*.*")]
            )
            tk_root.destroy()
            if not file_selection:
                print("Файлы не выбраны. Выход.")
                return
            # Process the entire directory of the first selected file
            root = os.path.dirname(file_selection[0])
        except ImportError:
            print("tkinter not available, use: segy2journal.py <directory>")
            return
    else:
        root = os.path.abspath(root)
        if not os.path.isdir(root):
            print(f"Error: directory not found: {root}")
            sys.exit(1)

    files = scan_directory(root, recursive=args.recursive)

    total = len(files)
    print(f"Каталог: {root}")
    if total:
        print(f"Найдено файлов SEG-Y: {total}\n")
    else:
        print("SEG-Y файлы не найдены.")
        return

    results = process_directory(files, root, recursive=args.recursive)
    if not results:
        return

    # Output goes into the same directory
    out_name = args.output or "segy_journal.xlsx"
    if not os.path.isabs(out_name):
        out_name = os.path.join(root, out_name)
    write_journal_xlsx(results, os.path.abspath(out_name))


if __name__ == "__main__":
    main()